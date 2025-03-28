"""
Script documentation

- Tool parameters are accessed using arcpy.GetParameter() or 
                                     arcpy.GetParameterAsText()
- Update derived parameter values using arcpy.SetParameter() or
                                        arcpy.SetParameterAsText()
"""
#################################################################################################################################################################################
## Libraries
import os
import sys
import arcpy
import openpyxl
import pandas as pd
from datetime import datetime
import itertools

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
#################################################################################################################################################################################
## Functions

def adjustTime(time_string:str, time_format='%Y-%m-%dT%H:%M:%S.%f')->str:
    """
    The Time Format returned from ArcGIS is '%Y-%m-%dT%H:%M:%S.%f'. As long as the timestamp is returned from the arcpy.Describe().dateModified/Accessed/Created, it will be this fromat.
    Args: time_zone
    """
    utc_datetime = datetime.strptime(time_string, time_format)
    local_datetime = utc_datetime.astimezone()
    local_date_str = local_datetime.strftime("%Y/%m/%d")
    
    return local_date_str

def checkGdb(gdb_path, workspace):
    ## Simply makes sure that the workspace is set to the correct FGDB. The program stops if it doesn't match
    if gdb_path != workspace:
        sys.exit("!!! Workspace and FGDB Path Dont Match !!!")
    

def generateFeatureClassList()->list:
    datasets = arcpy.ListDatasets("*", "Feature")
    feature_class_list = []
    if datasets:
        for dataset in datasets:
            for feature_class in arcpy.ListFeatureClasses(wild_card="*", feature_dataset=dataset):
                feature_class_list.append(os.path.join(arcpy.env.workspace, dataset, feature_class))

    for i in arcpy.ListFeatureClasses("*"):
        feature_class_list.append(os.path.join(arcpy.env.workspace, i))
        
    return feature_class_list
    

def script_tool(gdb_list, excel_report):
    """Script code goes below"""
    df_dicts = []
    
    for gdb in gdb_list:
        arcpy.env.workspace = gdb
        checkGdb(gdb, arcpy.env.workspace)
        arcpy.AddMessage(f"File GDB: {gdb}")
        arcpy.AddMessage(f"Fetching Raster/Table/Feature Class Lists...")
        tables = [os.path.join(gdb, t) for t in arcpy.ListTables()]
        rasters = [os.path.join(gdb, r) for r in arcpy.ListRasters()]
        vectors = generateFeatureClassList()
        items = list(itertools.chain(tables, rasters, vectors))

        arcpy.AddMessage(f"Iterating Items...")     
        for item in items:
            desc = arcpy.Describe(item)
            temp_dict = {"GDB Name":gdb.split("\\")[-1]}
            temp_dict["Data Type"] = desc.dataType
            temp_dict["Item Name"] = desc.name
            #temp_dict["Created Date"] = adjustTime(desc.dateCreated) if desc.dataType != "RasterDataset" else "N/A" ## For some reason all of the dates are being returned as modified date
            temp_dict["Last Modified Date"] = adjustTime(desc.dateModified) if desc.dataType != "RasterDataset" else "N/A"
            temp_dict["WKID"] = f"{desc.spatialReference.factoryCode}" if desc.dataType != "Table" else "N/A"
            temp_dict["Spatial Reference"] = f"{desc.spatialReference.name}" if desc.dataType != "Table" else "N/A"
            temp_dict["Full Path"] = item
            
            df_dicts.append(temp_dict)

    arcpy.AddMessage(f"Generating Data Frame...")
    #df = pd.DataFrame(df_dicts, columns=["Item Name", "Data Type", "GDB Name", "Created Date", "Last Modified Date", "WKID","Spatial Reference", "Full Path"])
    df = pd.DataFrame(df_dicts, columns=["Item Name", "Data Type", "GDB Name", "Last Modified Date", "WKID","Spatial Reference", "Full Path"])
    arcpy.AddMessage(f"Exporting Data Frame...")
    df.to_excel(excel_report, index = False)

    arcpy.AddMessage(f"Formatting Excel Report...")
    wb = openpyxl.load_workbook(excel_report) ## Creates the Workbook Object

    sheet_name = wb.sheetnames[0] ## Creates a list of strings of the sheet names in the workbook

    ws = wb[sheet_name] ## Creates the Work Sheet object

    ## this conditional checks if the table already exists. If the table exists it is deleted and a new table is created.
    if sheet_name in ws.tables:
        del ws.tables[sheet_name]

    table_dimensions = ws.calculate_dimension() ## grabs the dimensions. these are used when creating the Table Object

    tbl = Table(displayName="GeodatabaseItems", ref=table_dimensions) ## Creates the Table object using the grabbed dimensions

    style = TableStyleInfo(name=f"TableStyleMedium{2}",showFirstColumn=True, showLastColumn=False, showRowStripes=True, showColumnStripes=False) ## Formats the table. This is where the count comes in.

    tbl.tableStyleInfo = style ## Applies the style info to the Table Object
    ws.add_table(tbl) ## Adds the Table Object to the Worksheet object

    ## the below deals with the Column Width and Wrapping
    d = {"A":35, "B":15, "C":25, "D":20, "E":15, "F":40, "G":150}
    for c in d:
        ws.column_dimensions[c].width = d[c]


    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)




    wb.save(excel_report)
    del wb
                
    return df 


if __name__ == "__main__":
    gdbs = arcpy.GetParameterAsText(0)
    excel_report = arcpy.GetParameterAsText(1)
    gdb_list = [g.replace("'", "") for g in gdbs.split(";")]

    arcpy.AddMessage(f'GDB List: {gdb_list}')
    arcpy.AddMessage(f"Report Path: {excel_report}")
    arcpy.AddMessage(f"--"*100)

    script_tool(gdb_list, excel_report)